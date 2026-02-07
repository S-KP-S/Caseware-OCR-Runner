import argparse
import base64
import csv
import hashlib
import json
import os
import re
import sys
import time
from datetime import datetime

import fitz  # PyMuPDF
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config import load_accounts_map, load_vendor_map


PROMPT = (
    "You are an OCR extraction assistant for accounting imports. Extract transactions from the provided "
    "document images. Return a single JSON object only with key: transactions. "
    "transactions must be a list of objects with keys: date, description, amount, direction, "
    "transaction_type, currency, page_number. "
    "date must be YYYY-MM-DD. amount should be the numeric value as shown (commas/decimals ok); do not make negative. "
    "direction must be 'Debit' or 'Credit'. "
    "transaction_type must be one of Deposit, Withdrawal, Fee, Interest, Refund, Transfer, Payment, Purchase, Tax, Other. "
    "currency must be ISO 4217 (e.g., CAD, USD). page_number is 1-based. "
    "For receipts/invoices without line items, create a single transaction using the total amount. "
    "For bank/credit card statements, return one transaction per line item. "
    "Do not include balances or subtotals. "
    "If you cannot find transactions, return an empty list. Do not include any other keys or wrap in markdown."
)

AMOUNT_PROMPT = (
    "You are correcting transaction amounts for accounting import. "
    "Use the provided document images to verify and fix ONLY the Amount values. "
    "You are given a JSON list of transactions with an index. "
    "Return a single JSON object with key: transactions, a list of objects with keys: index, amount. "
    "Do not change the number of items or their index. "
    "amount must be the numeric value as shown (commas/decimals ok) and must not be negative. "
    "If the amount is not visible, return an empty string for amount. "
    "Do not include any other keys or wrap in markdown."
)

EXPORT_CASEWARE = "caseware"
EXPORT_FULL = "full"
EXPORT_QBO = "qbo"
EXPORT_XERO = "xero"
EXPORT_FORMATS = {EXPORT_CASEWARE, EXPORT_FULL, EXPORT_QBO, EXPORT_XERO}

CASEWARE_FIELDNAMES = [
    "Date",
    "Description",
    "Amount",
    "Direction",
    "Transaction Type",
    "Currency",
    "Source-File Name",
    "Page Number",
]
FULL_FIELDNAMES = CASEWARE_FIELDNAMES + ["Vendor", "Account", "Confidence", "Flags"]
QBO_FIELDNAMES = ["Date", "Description", "Amount"]
XERO_FIELDNAMES = ["Date", "Amount", "Payee", "Description", "Reference", "Transaction Type"]

SUMMARY_SHEET_TOTALS = "Monthly Totals"
SUMMARY_SHEET_MISSING = "Missing Fields"
SUMMARY_SHEET_DUPLICATES = "Potential Duplicates"
SUMMARY_SHEET_LOW_CONF = "Low Confidence"

OUTPUT_DIRNAME = "Extracted-Data"
CACHE_DIRNAME = ".ocr_cache"

DEFAULT_CURRENCY = "CAD"
DEFAULT_MODEL = "nvidia/nemotron-nano-12b-v2-vl:free"
DEFAULT_RPM = 12
DEFAULT_MAX_RETRIES = 5
DEFAULT_RETRY_BACKOFF = 5
DEFAULT_RETRY_MAX_SLEEP = 60
DEFAULT_BATCH_SIZE = 4

LOW_CONFIDENCE_THRESHOLD = 70
DUPLICATE_FLAG_PREFIX = "Duplicate of row"

KEY_FIELDS = [
    "Date",
    "Description",
    "Amount",
    "Direction",
    "Transaction Type",
    "Currency",
    "Source-File Name",
]

SKIP_EXTENSIONS = {".xls", ".xlsx", ".xlsm", ".xlsb"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}
CSV_EXTENSIONS = {".csv"}

TRANSACTION_TYPE_SET = {
    "Deposit",
    "Withdrawal",
    "Fee",
    "Interest",
    "Refund",
    "Transfer",
    "Payment",
    "Purchase",
    "Tax",
    "Other",
}

CSV_DATE_KEYS = {"date", "transactiondate", "postingdate", "postdate", "transdate", "valuedate"}
CSV_DESC_KEYS = {"description", "details", "memo", "payee", "merchant", "name", "narration", "reference"}
CSV_AMOUNT_KEYS = {"amount", "transactionamount", "amt", "value", "total", "amountcad", "amountcredit", "amountdebit"}
CSV_DEBIT_KEYS = {"debit", "withdrawal", "withdrawals", "moneyout", "debitamount", "charges", "charge", "paidout"}
CSV_CREDIT_KEYS = {"credit", "deposit", "deposits", "moneyin", "creditamount", "paymentreceived", "receipt"}
CSV_DIRECTION_KEYS = {"direction", "drcr", "debitcredit", "debitcreditindicator"}
CSV_CURRENCY_KEYS = {"currency", "curr", "ccy", "currencycode", "iso"}
CSV_TYPE_KEYS = {"transactiontype", "type", "category", "classification", "transtype", "transaction_type"}

RATE_LIMITER = None
PROGRESS_CALLBACK = None


class ProcessingStopped(RuntimeError):
    pass


# SECTION: utilities
def set_progress_callback(callback):
    global PROGRESS_CALLBACK
    PROGRESS_CALLBACK = callback


def _emit_progress(message, current=None, total=None, level="INFO"):
    event = {"message": str(message), "current": current, "total": total, "level": level}
    if PROGRESS_CALLBACK:
        try:
            PROGRESS_CALLBACK(event)
            return
        except TypeError:
            PROGRESS_CALLBACK(message, current, total, level)
            return
        except Exception:
            pass
    print(f"{level}: {message}")


def _check_stop(stop_check):
    if callable(stop_check) and stop_check():
        raise ProcessingStopped("Stopped by user.")


def parse_args():
    parser = argparse.ArgumentParser(
        description="OCR PDFs/images/CSVs via OpenRouter and write accounting-friendly CSV exports plus summary."
    )
    parser.add_argument("--input", required=True, help="Path to a PDF/image/CSV or a directory")
    parser.add_argument("--output", help="Path to output CSV (single or batch)")
    parser.add_argument("--model", default=DEFAULT_MODEL, help="OpenRouter model")
    parser.add_argument("--max-pages", type=int, default=12, help="Max pages to render")
    parser.add_argument("--zoom", type=float, default=2.0, help="Render zoom factor")
    parser.add_argument("--currency", default=DEFAULT_CURRENCY, help="Default currency code")
    parser.add_argument("--rpm", type=int, default=DEFAULT_RPM, help="Max OpenRouter requests per minute")
    parser.add_argument("--max-retries", type=int, default=DEFAULT_MAX_RETRIES, help="Max OpenRouter retries")
    parser.add_argument("--retry-backoff", type=int, default=DEFAULT_RETRY_BACKOFF, help="Base retry backoff seconds")
    parser.add_argument("--retry-max-sleep", type=int, default=DEFAULT_RETRY_MAX_SLEEP, help="Max retry sleep seconds")
    parser.add_argument("--recursive", action="store_true", help="Process subfolders (default)")
    parser.add_argument("--no-recursive", dest="recursive", action="store_false", help="Disable subfolder processing")
    parser.add_argument("--no-refine-amounts", action="store_false", dest="refine_amounts", help="Disable amount correction")
    parser.add_argument("--no-cache", action="store_false", dest="use_cache", help="Disable file cache")
    parser.add_argument(
        "--export-format",
        default=EXPORT_CASEWARE,
        choices=sorted(EXPORT_FORMATS),
        help="CSV export format",
    )
    parser.set_defaults(refine_amounts=True, recursive=True, use_cache=True)
    return parser.parse_args()


def safe_str(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)
    return str(value)


def normalize_header(value):
    text = safe_str(value).lower().strip()
    return re.sub(r"[^a-z0-9]", "", text)


def normalize_amount_text(value):
    text = safe_str(value).strip()
    if not text:
        return ""
    if text.startswith("(") and text.endswith(")"):
        text = text[1:-1].strip()
    text = re.sub(r"^-\s*", "", text)
    return text


def strip_code_fences(text):
    text = text.strip()
    if not text.startswith("```"):
        return text
    lines = text.splitlines()
    if lines and lines[0].startswith("```"):
        lines = lines[1:]
    if lines and lines[-1].startswith("```"):
        lines = lines[:-1]
    return "\n".join(lines).strip()


def parse_json(text):
    text = strip_code_fences(text)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start == -1 or end == -1 or end <= start:
            raise
        return json.loads(text[start:end + 1])


class RateLimiter:
    def __init__(self, rpm):
        self.min_interval = 60.0 / rpm if rpm and rpm > 0 else 0.0
        self.last_time = 0.0

    def wait(self):
        if self.min_interval <= 0:
            return
        now = time.time()
        elapsed = now - self.last_time
        if elapsed < self.min_interval:
            time.sleep(self.min_interval - elapsed)
        self.last_time = time.time()


def is_skipped_extension(path):
    return os.path.splitext(path)[1].lower() in SKIP_EXTENSIONS


def is_csv_extension(path):
    return os.path.splitext(path)[1].lower() in CSV_EXTENSIONS


def is_image_extension(path):
    return os.path.splitext(path)[1].lower() in IMAGE_EXTENSIONS


# SECTION: model calls
def render_pdf_to_png_bytes(pdf_path, max_pages, zoom):
    doc = fitz.open(pdf_path)
    images = []
    for i, page in enumerate(doc):
        if max_pages and i >= max_pages:
            break
        matrix = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=matrix, alpha=False)
        images.append(pix.tobytes("png"))
    doc.close()
    return images, len(images)


def load_images(path, max_pages, zoom):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        return render_pdf_to_png_bytes(path, max_pages, zoom)
    if ext in IMAGE_EXTENSIONS:
        with open(path, "rb") as f:
            return [f.read()], 1
    return None, 0


def build_content(images, prompt_text):
    content = [{"type": "text", "text": prompt_text}]
    for img in images:
        b64 = base64.b64encode(img).decode("ascii")
        content.append({"type": "image_url", "image_url": {"url": "data:image/png;base64," + b64}})
    return content


def compute_retry_sleep(response, attempt, backoff_base, backoff_max):
    reset_header = response.headers.get("X-RateLimit-Reset")
    if reset_header:
        try:
            reset_val = float(reset_header)
            if reset_val > 1e12:
                reset_val /= 1000.0
            delay = max(0.0, reset_val - time.time()) + 1.0
            return min(backoff_max, delay)
        except ValueError:
            pass
    delay = backoff_base * (2 ** (attempt - 1))
    return min(backoff_max, delay)


def call_openrouter(content, api_key, model, max_retries, backoff_base, backoff_max):
    if not api_key:
        raise RuntimeError("OPENROUTER_API_KEY is not set.")

    payload = {"model": model, "temperature": 0, "messages": [{"role": "user", "content": content}]}
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "X-Title": "Caseware OCR Runner",
    }

    for attempt in range(1, max_retries + 1):
        if RATE_LIMITER:
            RATE_LIMITER.wait()
        response = requests.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers=headers,
            json=payload,
            timeout=120,
        )

        if response.status_code == 200:
            try:
                data = response.json()
                content_text = data["choices"][0]["message"]["content"]
                return parse_json(content_text)
            except Exception as exc:
                if attempt < max_retries:
                    sleep_time = min(backoff_max, backoff_base * (2 ** (attempt - 1)))
                    _emit_progress(f"OpenRouter parse error, retrying in {sleep_time:.1f}s", level="WARN")
                    time.sleep(sleep_time)
                    continue
                raise RuntimeError(f"OpenRouter response parse failed: {exc}") from exc

        retryable = response.status_code in (408, 409, 429, 500, 502, 503, 504)
        if "degraded" in response.text.lower():
            retryable = True
        if retryable and attempt < max_retries:
            sleep_time = compute_retry_sleep(response, attempt, backoff_base, backoff_max)
            _emit_progress(
                f"OpenRouter retry {attempt}/{max_retries} in {sleep_time:.1f}s (status {response.status_code})",
                level="WARN",
            )
            time.sleep(sleep_time)
            continue

        raise RuntimeError(f"OpenRouter request failed: {response.status_code} {response.text}")

    raise RuntimeError("OpenRouter request failed after retries")


def normalize_transactions(extracted):
    if isinstance(extracted, list):
        return extracted
    if isinstance(extracted, dict):
        txns = extracted.get("transactions")
        if isinstance(txns, list):
            return txns
    return []


def _offset_page_number(item, offset):
    if not isinstance(item, dict) or offset <= 0:
        return
    raw = item.get("page_number")
    if raw is None:
        raw = item.get("page")
    if raw is None:
        return
    match = re.search(r"\d+", safe_str(raw))
    if not match:
        return
    item["page_number"] = str(int(match.group(0)) + offset)


def call_openrouter_batched(
    images,
    api_key,
    model,
    max_retries,
    backoff_base,
    backoff_max,
    prompt_text=PROMPT,
    batch_size=DEFAULT_BATCH_SIZE,
    stop_check=None,
):
    if not images:
        return {"transactions": []}
    if len(images) <= batch_size:
        return call_openrouter(build_content(images, prompt_text), api_key, model, max_retries, backoff_base, backoff_max)

    merged = []
    batch_count = (len(images) + batch_size - 1) // batch_size
    for i in range(batch_count):
        _check_stop(stop_check)
        start = i * batch_size
        end = min(start + batch_size, len(images))
        _emit_progress(f"OCR batch {i + 1}/{batch_count} (pages {start + 1}-{end})")
        extracted = call_openrouter(
            build_content(images[start:end], prompt_text),
            api_key,
            model,
            max_retries,
            backoff_base,
            backoff_max,
        )
        for item in normalize_transactions(extracted):
            if not isinstance(item, dict):
                continue
            item_copy = dict(item)
            _offset_page_number(item_copy, start)
            merged.append(item_copy)
    return {"transactions": merged}


# SECTION: transforms
def normalize_transaction_type(value):
    if not value:
        return ""
    cleaned = str(value).strip()
    lowered = cleaned.lower()
    alias_map = {
        "credit": "Deposit",
        "cr": "Deposit",
        "deposit": "Deposit",
        "debit": "Withdrawal",
        "dr": "Withdrawal",
        "withdrawal": "Withdrawal",
        "withdraw": "Withdrawal",
        "fee": "Fee",
        "charge": "Fee",
        "interest": "Interest",
        "refund": "Refund",
        "reversal": "Refund",
        "transfer": "Transfer",
        "payment": "Payment",
        "purchase": "Purchase",
        "pos": "Purchase",
        "tax": "Tax",
    }
    if lowered in alias_map:
        return alias_map[lowered]
    cleaned_title = cleaned.title()
    if cleaned_title in TRANSACTION_TYPE_SET:
        return cleaned_title
    return cleaned_title if cleaned_title else ""


def infer_direction(direction, transaction_type):
    if direction:
        cleaned = str(direction).strip().lower()
        if cleaned in ("debit", "dr") or "debit" in cleaned or "withdraw" in cleaned:
            return "Debit"
        if cleaned in ("credit", "cr") or "credit" in cleaned or "deposit" in cleaned:
            return "Credit"
    if transaction_type:
        ttype = str(transaction_type).strip().lower()
        if ttype in ("deposit", "interest", "refund"):
            return "Credit"
        if ttype in ("withdrawal", "fee", "payment", "purchase", "tax"):
            return "Debit"
    return ""


def infer_transaction_type_from_description(description):
    if not description:
        return ""
    text = str(description).lower()
    if "refund" in text or "reversal" in text:
        return "Refund"
    if "interest" in text:
        return "Interest"
    if "fee" in text or "charge" in text:
        return "Fee"
    if "tax" in text or "hst" in text or "cra" in text:
        return "Tax"
    if "transfer" in text or "e-transfer" in text or "etransfer" in text:
        return "Transfer"
    if "payroll" in text or "payment" in text:
        return "Payment"
    if "deposit" in text or "payment received" in text:
        return "Deposit"
    if "withdrawal" in text or "atm" in text:
        return "Withdrawal"
    if "purchase" in text or "pos" in text or "card" in text:
        return "Purchase"
    return ""


def parse_date_info(value):
    if not value:
        return None, "missing"
    text = str(value).strip()
    match = re.search(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", text)
    if match:
        year, month, day = match.groups()
        try:
            return datetime(int(year), int(month), int(day)).date(), "ok"
        except ValueError:
            return None, "invalid"
    match = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})", text)
    if match:
        month, day, year = match.groups()
        try:
            return datetime(int(year), int(month), int(day)).date(), "ok"
        except ValueError:
            return None, "invalid"
    return None, "unparseable"


def parse_amount(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    negative_by_parens = text.startswith("(") and text.endswith(")")
    if negative_by_parens:
        text = text[1:-1].strip()
    cleaned = re.sub(r"[^\d,.\-]", "", text)
    if not cleaned:
        return None
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    elif "," in cleaned and "." not in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        amount = float(cleaned)
        if negative_by_parens and amount > 0:
            amount = -amount
        return amount
    except ValueError:
        return None


def _get_first_value(row, keys):
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _empty_row():
    return {
        "Date": "",
        "Description": "",
        "Amount": "",
        "Direction": "",
        "Transaction Type": "",
        "Currency": "",
        "Source-File Name": "",
        "Page Number": "",
        "Vendor": "",
        "Account": "",
        "Confidence": "",
        "Flags": "",
    }


def ensure_row_schema(row):
    normalized = _empty_row()
    if isinstance(row, dict):
        for key in normalized.keys():
            if key in row:
                normalized[key] = row.get(key, "")
    return normalized


def build_rows(input_path, extracted, default_currency, page_count, source_name=None):
    source_name = source_name or os.path.basename(input_path)
    rows = []
    for item in normalize_transactions(extracted):
        if not isinstance(item, dict):
            continue
        date = safe_str(item.get("date") or item.get("transaction_date"))
        description = safe_str(item.get("description") or item.get("memo") or item.get("payee") or item.get("vendor") or item.get("details"))
        amount = safe_str(item.get("amount") or item.get("total") or item.get("value"))
        transaction_type = normalize_transaction_type(item.get("transaction_type") or item.get("type"))
        if not transaction_type:
            transaction_type = infer_transaction_type_from_description(description)
        direction = infer_direction(item.get("direction") or item.get("sign"), transaction_type)
        amount_value = parse_amount(amount)
        if amount_value is not None and amount_value < 0 and not direction:
            direction = "Debit"
        currency = safe_str(item.get("currency") or default_currency)
        page_number = safe_str(item.get("page_number") or item.get("page") or "")
        if not page_number and page_count == 1:
            page_number = "1"
        row = _empty_row()
        row.update(
            {
                "Date": date,
                "Description": description,
                "Amount": normalize_amount_text(amount),
                "Direction": direction,
                "Transaction Type": transaction_type,
                "Currency": currency,
                "Source-File Name": source_name,
                "Page Number": page_number,
            }
        )
        rows.append(row)
    return rows


# SECTION: csv parse and enrich
def parse_csv_file(path, default_currency, source_name):
    rows = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        raw_rows = list(csv.reader(f, dialect))

    if not raw_rows:
        return rows

    norm_headers = [normalize_header(h) for h in raw_rows[0]]
    for raw in raw_rows[1:]:
        if not any(cell.strip() for cell in raw if isinstance(cell, str)):
            continue
        row_dict = {}
        for idx, key in enumerate(norm_headers):
            value = raw[idx] if idx < len(raw) else ""
            row_dict[key] = value.strip() if isinstance(value, str) else str(value)

        date = _get_first_value(row_dict, CSV_DATE_KEYS)
        description = _get_first_value(row_dict, CSV_DESC_KEYS)
        amount = _get_first_value(row_dict, CSV_AMOUNT_KEYS)
        debit = _get_first_value(row_dict, CSV_DEBIT_KEYS)
        credit = _get_first_value(row_dict, CSV_CREDIT_KEYS)
        direction_raw = _get_first_value(row_dict, CSV_DIRECTION_KEYS)
        type_raw = _get_first_value(row_dict, CSV_TYPE_KEYS)
        currency = _get_first_value(row_dict, CSV_CURRENCY_KEYS) or default_currency
        if not any([date, description, amount, debit, credit]):
            continue

        transaction_type = normalize_transaction_type(type_raw)
        direction = infer_direction(direction_raw, transaction_type)
        if not direction and type_raw:
            maybe_direction = infer_direction(type_raw, transaction_type)
            if maybe_direction:
                direction = maybe_direction
            elif not transaction_type:
                transaction_type = normalize_transaction_type(type_raw)

        if debit and not amount:
            amount = debit
            direction = direction or "Debit"
        elif credit and not amount:
            amount = credit
            direction = direction or "Credit"

        if amount:
            amount_value = parse_amount(amount)
            if amount_value is not None:
                if amount_value < 0 and not direction:
                    direction = "Debit"
                elif amount_value > 0 and not direction and not (debit or credit):
                    direction = "Credit"
            amount = normalize_amount_text(amount)

        if not transaction_type:
            transaction_type = infer_transaction_type_from_description(description)

        row = _empty_row()
        row.update(
            {
                "Date": date,
                "Description": description,
                "Amount": amount,
                "Direction": direction,
                "Transaction Type": transaction_type,
                "Currency": currency,
                "Source-File Name": source_name,
            }
        )
        rows.append(row)
    return rows


def refine_amounts(rows, images, api_key, model, max_retries, backoff_base, backoff_max):
    payload_rows = []
    for idx, row in enumerate(rows):
        payload_rows.append(
            {
                "index": idx,
                "date": row.get("Date", ""),
                "description": row.get("Description", ""),
                "direction": row.get("Direction", ""),
                "transaction_type": row.get("Transaction Type", ""),
                "currency": row.get("Currency", ""),
                "page_number": row.get("Page Number", ""),
                "amount": row.get("Amount", ""),
            }
        )
    prompt_text = AMOUNT_PROMPT + "\nTransactions JSON:\n" + json.dumps(payload_rows, ensure_ascii=True)
    extracted = call_openrouter(build_content(images, prompt_text), api_key, model, max_retries, backoff_base, backoff_max)

    if isinstance(extracted, list):
        corrections = extracted
    elif isinstance(extracted, dict):
        corrections = extracted.get("transactions") or extracted.get("amounts") or []
    else:
        corrections = []

    updated = 0
    for item in corrections:
        if not isinstance(item, dict):
            continue
        idx = item.get("index")
        try:
            idx = int(idx)
        except (TypeError, ValueError):
            continue
        if idx < 0 or idx >= len(rows):
            continue
        amount = normalize_amount_text(item.get("amount"))
        if amount:
            rows[idx]["Amount"] = amount
            updated += 1
    return updated


def _normalize_flags(value):
    parts = [p.strip() for p in safe_str(value).split(";")]
    parts = [p for p in parts if p]
    deduped = []
    seen = set()
    for part in parts:
        token = part.lower()
        if token in seen:
            continue
        seen.add(token)
        deduped.append(part)
    return deduped


def _set_flags(row, flags):
    row["Flags"] = "; ".join(flags)


def _append_flag(row, flag):
    flags = _normalize_flags(row.get("Flags", ""))
    token = safe_str(flag).strip()
    if not token:
        return
    if token.lower() not in {f.lower() for f in flags}:
        flags.append(token)
    _set_flags(row, flags)


def score_confidence(row):
    score = 100
    flags = []

    date_value = safe_str(row.get("Date")).strip()
    parsed_date, date_status = parse_date_info(date_value)
    if not date_value:
        score -= 30
        flags.append("Missing date")
    elif date_status != "ok":
        score -= 20
        flags.append("Unparseable date")
    else:
        today = datetime.now().date()
        if parsed_date > today:
            score -= 10
            flags.append("Future date")
        if parsed_date.year < 2000:
            score -= 10
            flags.append("Pre-2000 date")

    amount_text = safe_str(row.get("Amount")).strip()
    amount_value = parse_amount(amount_text)
    if not amount_text:
        score -= 30
        flags.append("Missing amount")
    elif amount_value is None:
        score -= 30
        flags.append("Unparseable amount")
    else:
        abs_amount = abs(amount_value)
        if abs_amount == 0:
            score -= 10
            flags.append("Zero amount")
        if abs_amount > 1_000_000:
            score -= 5
            flags.append("Large amount")

    description = safe_str(row.get("Description")).strip()
    if not description:
        score -= 15
        flags.append("Missing description")
    elif len(description) < 4:
        score -= 5
        flags.append("Short description")

    direction = safe_str(row.get("Direction")).strip().title()
    if direction not in ("Debit", "Credit"):
        score -= 10
        flags.append("Invalid direction")

    ttype = normalize_transaction_type(row.get("Transaction Type"))
    if not ttype or ttype not in TRANSACTION_TYPE_SET:
        score -= 5
        flags.append("Invalid transaction type")

    return max(0, min(100, int(score))), flags


def normalize_vendor(description, vendor_map):
    if not description or not isinstance(vendor_map, dict):
        return ""
    desc = safe_str(description).upper()
    ranked = sorted(vendor_map.items(), key=lambda x: len(safe_str(x[0]).strip()), reverse=True)
    for pattern, vendor in ranked:
        pat = safe_str(pattern).strip()
        if pat and pat.upper() in desc:
            return safe_str(vendor).strip()
    return ""


def lookup_account(vendor, accounts_map):
    if not vendor or not isinstance(accounts_map, dict):
        return ""
    vendor_text = safe_str(vendor).strip()
    if vendor_text in accounts_map:
        return safe_str(accounts_map[vendor_text]).strip()
    needle = vendor_text.lower()
    for key, value in accounts_map.items():
        if safe_str(key).strip().lower() == needle:
            return safe_str(value).strip()
    return ""


def enrich_rows(rows, vendor_map, accounts_map):
    enriched = []
    for row in rows:
        norm = ensure_row_schema(row)
        norm["Flags"] = ""
        norm["Transaction Type"] = normalize_transaction_type(norm.get("Transaction Type"))
        if not norm["Transaction Type"]:
            norm["Transaction Type"] = infer_transaction_type_from_description(norm.get("Description", ""))
        if not norm["Direction"]:
            norm["Direction"] = infer_direction("", norm["Transaction Type"])
        if not norm["Currency"]:
            norm["Currency"] = DEFAULT_CURRENCY
        norm["Vendor"] = normalize_vendor(norm.get("Description", ""), vendor_map)
        norm["Account"] = lookup_account(norm["Vendor"], accounts_map)
        score, flags = score_confidence(norm)
        norm["Confidence"] = score
        for flag in flags:
            _append_flag(norm, flag)
        enriched.append(norm)
    return enriched


def _duplicate_key(row):
    date_key = safe_str(row.get("Date")).strip().lower()
    amount = parse_amount(row.get("Amount"))
    amount_key = "" if amount is None else f"{abs(amount):.2f}"
    desc_key = safe_str(row.get("Description")).strip().lower()[:30]
    return date_key, amount_key, desc_key


def detect_duplicates(rows):
    for row in rows:
        flags = [f for f in _normalize_flags(row.get("Flags", "")) if not f.lower().startswith(DUPLICATE_FLAG_PREFIX.lower())]
        _set_flags(row, flags)

    seen = {}
    duplicates = 0
    for idx, row in enumerate(rows, start=1):
        key = _duplicate_key(row)
        if not any(key):
            continue
        if key in seen:
            duplicates += 1
            _append_flag(row, f"{DUPLICATE_FLAG_PREFIX} {seen[key]}")
        else:
            seen[key] = idx
    return duplicates


# SECTION: export and summary
def _get_export_fieldnames(export_format):
    if export_format == EXPORT_CASEWARE:
        return CASEWARE_FIELDNAMES
    if export_format == EXPORT_FULL:
        return FULL_FIELDNAMES
    if export_format == EXPORT_QBO:
        return QBO_FIELDNAMES
    if export_format == EXPORT_XERO:
        return XERO_FIELDNAMES
    raise ValueError(f"Unsupported export format: {export_format}")


def _format_row_for_export(row, export_format):
    if export_format == EXPORT_CASEWARE:
        return {field: row.get(field, "") for field in CASEWARE_FIELDNAMES}
    if export_format == EXPORT_FULL:
        return {field: row.get(field, "") for field in FULL_FIELDNAMES}
    if export_format == EXPORT_QBO:
        return {"Date": row.get("Date", ""), "Description": row.get("Description", ""), "Amount": row.get("Amount", "")}
    if export_format == EXPORT_XERO:
        source = safe_str(row.get("Source-File Name", ""))
        page = safe_str(row.get("Page Number", "")).strip()
        reference = source if not page else f"{source} p{page}"
        payee = safe_str(row.get("Vendor", "")).strip() or safe_str(row.get("Description", "")).strip()[:80]
        return {
            "Date": row.get("Date", ""),
            "Amount": row.get("Amount", ""),
            "Payee": payee,
            "Description": row.get("Description", ""),
            "Reference": reference,
            "Transaction Type": row.get("Transaction Type", ""),
        }
    raise ValueError(f"Unsupported export format: {export_format}")


def write_csv(output_path, rows, export_format=EXPORT_CASEWARE):
    fieldnames = _get_export_fieldnames(export_format)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(_format_row_for_export(row, export_format))


def _style_header(ws, headers):
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def _auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(safe_str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 70)


def write_summary(output_path, rows):
    wb = Workbook()
    totals_ws = wb.active
    totals_ws.title = SUMMARY_SHEET_TOTALS
    _style_header(totals_ws, ["Month", "Row Count", "Debit Total", "Credit Total", "Net Total"])

    monthly = {}
    missing_rows = []

    for row in rows:
        missing = []
        for field in KEY_FIELDS:
            value = row.get(field)
            if value is None or str(value).strip() == "":
                missing.append(field)

        parsed_date, date_status = parse_date_info(row.get("Date"))
        if row.get("Date") and date_status != "ok":
            missing.append("Date (invalid)" if date_status == "invalid" else "Date (unparseable)")

        amount_value = parse_amount(row.get("Amount"))
        if row.get("Amount") and amount_value is None:
            missing.append("Amount (unparseable)")

        if missing:
            missing_rows.append((row, ", ".join(sorted(set(missing)))))

        if not parsed_date or amount_value is None:
            continue

        month_key = f"{parsed_date.year:04d}-{parsed_date.month:02d}"
        bucket = monthly.setdefault(month_key, {"count": 0, "debit": 0.0, "credit": 0.0, "net": 0.0})
        bucket["count"] += 1
        direction = str(row.get("Direction", "")).strip().title()
        abs_amount = abs(amount_value)
        if direction == "Credit":
            bucket["credit"] += abs_amount
            bucket["net"] += abs_amount
        elif direction == "Debit":
            bucket["debit"] += abs_amount
            bucket["net"] -= abs_amount
        else:
            bucket["net"] += amount_value

    for month_key in sorted(monthly.keys()):
        bucket = monthly[month_key]
        totals_ws.append([month_key, bucket["count"], round(bucket["debit"], 2), round(bucket["credit"], 2), round(bucket["net"], 2)])

    missing_ws = wb.create_sheet(title=SUMMARY_SHEET_MISSING)
    _style_header(missing_ws, FULL_FIELDNAMES + ["Missing Fields"])
    for row, missing in missing_rows:
        missing_ws.append([row.get(field, "") for field in FULL_FIELDNAMES] + [missing])

    dup_ws = wb.create_sheet(title=SUMMARY_SHEET_DUPLICATES)
    _style_header(dup_ws, FULL_FIELDNAMES)
    for row in rows:
        if DUPLICATE_FLAG_PREFIX.lower() in safe_str(row.get("Flags", "")).lower():
            dup_ws.append([row.get(field, "") for field in FULL_FIELDNAMES])

    low_ws = wb.create_sheet(title=SUMMARY_SHEET_LOW_CONF)
    _style_header(low_ws, FULL_FIELDNAMES)
    for row in rows:
        try:
            conf = int(float(row.get("Confidence", 0)))
        except (TypeError, ValueError):
            conf = 0
        if conf < LOW_CONFIDENCE_THRESHOLD:
            low_ws.append([row.get(field, "") for field in FULL_FIELDNAMES])

    for ws in wb.worksheets:
        _auto_fit_columns(ws)
    wb.save(output_path)


# SECTION: caching and pipeline
def get_source_name(root_dir, path):
    if root_dir:
        try:
            return os.path.relpath(path, root_dir)
        except ValueError:
            return os.path.basename(path)
    return os.path.basename(path)


def iter_input_files(input_path, recursive):
    if recursive:
        for root, dirs, files in os.walk(input_path):
            dirs[:] = [d for d in dirs if d != OUTPUT_DIRNAME and not d.startswith(".")]
            for name in files:
                yield os.path.join(root, name)
    else:
        for name in sorted(os.listdir(input_path)):
            yield os.path.join(input_path, name)


def compute_file_hash(path):
    digest = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def _ensure_cache_dir(output_dir):
    cache_dir = os.path.join(output_dir, CACHE_DIRNAME)
    os.makedirs(cache_dir, exist_ok=True)
    return cache_dir


def _cache_path(cache_dir, file_hash):
    return os.path.join(cache_dir, f"{file_hash}.json")


def _load_cached_rows(cache_file, source_name):
    if not os.path.isfile(cache_file):
        return None
    try:
        with open(cache_file, "r", encoding="utf-8") as f:
            payload = json.load(f)
        rows = payload.get("rows", [])
        if not isinstance(rows, list):
            return None
        normalized = []
        for row in rows:
            norm = ensure_row_schema(row)
            norm["Source-File Name"] = source_name
            normalized.append(norm)
        return normalized
    except Exception:
        return None


def _save_cached_rows(cache_file, rows):
    payload = {
        "cached_at": datetime.utcnow().isoformat() + "Z",
        "rows": rows,
    }
    with open(cache_file, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def _extract_rows_from_file(
    path,
    source_name,
    api_key,
    model,
    max_pages,
    zoom,
    currency,
    max_retries,
    retry_backoff,
    retry_max_sleep,
    refine_amounts_enabled,
    stop_check=None,
):
    _check_stop(stop_check)
    if is_csv_extension(path):
        return parse_csv_file(path, currency, source_name)

    if not (path.lower().endswith(".pdf") or is_image_extension(path)):
        return []

    images, page_count = load_images(path, max_pages, zoom)
    if not images:
        return []

    extracted = call_openrouter_batched(
        images,
        api_key,
        model,
        max_retries,
        retry_backoff,
        retry_max_sleep,
        prompt_text=PROMPT,
        batch_size=DEFAULT_BATCH_SIZE,
        stop_check=stop_check,
    )
    rows = build_rows(path, extracted, currency, page_count, source_name=source_name)
    if rows and refine_amounts_enabled:
        try:
            updated = refine_amounts(rows, images, api_key, model, max_retries, retry_backoff, retry_max_sleep)
            if updated:
                _emit_progress(f"{source_name} amounts refined: {updated}")
        except Exception as exc:
            _emit_progress(f"{source_name} amount refine failed: {exc}", level="WARN")
    return rows


def _process_file(
    path,
    source_name,
    output_dir,
    api_key,
    model,
    max_pages,
    zoom,
    currency,
    max_retries,
    retry_backoff,
    retry_max_sleep,
    refine_amounts_enabled,
    use_cache,
    vendor_map,
    accounts_map,
    stop_check=None,
):
    cache_dir = _ensure_cache_dir(output_dir)
    cache_file = _cache_path(cache_dir, compute_file_hash(path))

    rows = None
    if use_cache:
        rows = _load_cached_rows(cache_file, source_name)
        if rows is not None:
            _emit_progress(f"Cache hit: {source_name}")

    if rows is None:
        rows = _extract_rows_from_file(
            path,
            source_name,
            api_key,
            model,
            max_pages,
            zoom,
            currency,
            max_retries,
            retry_backoff,
            retry_max_sleep,
            refine_amounts_enabled,
            stop_check=stop_check,
        )
        if use_cache and rows:
            _save_cached_rows(cache_file, rows)

    rows = enrich_rows(rows, vendor_map, accounts_map)
    detect_duplicates(rows)
    return rows


def _resolve_output_paths_for_directory(input_path, output_path):
    output_dir = os.path.join(input_path, OUTPUT_DIRNAME)
    os.makedirs(output_dir, exist_ok=True)
    folder_name = os.path.basename(os.path.abspath(input_path))
    csv_path = output_path or os.path.join(output_dir, f"{folder_name}_transactions.csv")
    summary_path = os.path.join(output_dir, f"{folder_name}_summary.xlsx")
    return output_dir, csv_path, summary_path


def _resolve_output_paths_for_file(input_path, output_path):
    if output_path:
        csv_path = output_path
        summary_path = os.path.splitext(output_path)[0] + "_summary.xlsx"
        output_dir = os.path.dirname(os.path.abspath(csv_path))
        os.makedirs(output_dir, exist_ok=True)
        return output_dir, csv_path, summary_path
    base_dir = os.path.dirname(os.path.abspath(input_path))
    output_dir = os.path.join(base_dir, OUTPUT_DIRNAME)
    os.makedirs(output_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    csv_path = os.path.join(output_dir, f"{base_name}_transactions.csv")
    summary_path = os.path.join(output_dir, f"{base_name}_summary.xlsx")
    return output_dir, csv_path, summary_path


def process_directory(
    input_path,
    api_key,
    model=DEFAULT_MODEL,
    output_path=None,
    max_pages=12,
    zoom=2.0,
    currency=DEFAULT_CURRENCY,
    rpm=DEFAULT_RPM,
    max_retries=DEFAULT_MAX_RETRIES,
    retry_backoff=DEFAULT_RETRY_BACKOFF,
    retry_max_sleep=DEFAULT_RETRY_MAX_SLEEP,
    recursive=True,
    refine_amounts=True,
    use_cache=True,
    export_format=EXPORT_CASEWARE,
    vendor_map=None,
    accounts_map=None,
    stop_check=None,
):
    if not os.path.isdir(input_path):
        raise ValueError(f"Not a directory: {input_path}")
    if export_format not in EXPORT_FORMATS:
        raise ValueError(f"Invalid export format: {export_format}")

    global RATE_LIMITER
    RATE_LIMITER = RateLimiter(rpm)
    vendor_map = load_vendor_map() if vendor_map is None else vendor_map
    accounts_map = load_accounts_map() if accounts_map is None else accounts_map

    output_dir, csv_path, summary_path = _resolve_output_paths_for_directory(input_path, output_path)
    files = []
    for full_path in iter_input_files(input_path, recursive):
        if not os.path.isfile(full_path):
            continue
        if is_skipped_extension(full_path):
            continue
        if OUTPUT_DIRNAME in full_path.split(os.sep):
            continue
        if os.path.abspath(full_path) in {os.path.abspath(csv_path), os.path.abspath(summary_path)}:
            continue
        files.append(full_path)
    if not files:
        raise RuntimeError("No eligible files found in directory.")

    all_rows = []
    total = len(files)
    for idx, path in enumerate(files, start=1):
        _check_stop(stop_check)
        source_name = get_source_name(input_path, path)
        _emit_progress(f"Processing {idx}/{total}: {source_name}", current=idx, total=total)
        try:
            rows = _process_file(
                path=path,
                source_name=source_name,
                output_dir=output_dir,
                api_key=api_key,
                model=model,
                max_pages=max_pages,
                zoom=zoom,
                currency=currency,
                max_retries=max_retries,
                retry_backoff=retry_backoff,
                retry_max_sleep=retry_max_sleep,
                refine_amounts_enabled=refine_amounts,
                use_cache=use_cache,
                vendor_map=vendor_map,
                accounts_map=accounts_map,
                stop_check=stop_check,
            )
        except ProcessingStopped:
            raise
        except Exception as exc:
            _emit_progress(f"{source_name} failed: {exc}", current=idx, total=total, level="ERROR")
            continue
        if not rows:
            _emit_progress(f"No transactions found: {source_name}", current=idx, total=total, level="WARN")
            continue
        all_rows.extend(rows)

    if not all_rows:
        raise RuntimeError("No transactions were extracted.")

    detect_duplicates(all_rows)
    write_csv(csv_path, all_rows, export_format=export_format)
    write_summary(summary_path, all_rows)
    _emit_progress(f"Wrote {csv_path}")
    _emit_progress(f"Wrote {summary_path}")
    return all_rows, csv_path, summary_path


def process_single_file(
    input_path,
    api_key,
    model=DEFAULT_MODEL,
    output_path=None,
    max_pages=12,
    zoom=2.0,
    currency=DEFAULT_CURRENCY,
    rpm=DEFAULT_RPM,
    max_retries=DEFAULT_MAX_RETRIES,
    retry_backoff=DEFAULT_RETRY_BACKOFF,
    retry_max_sleep=DEFAULT_RETRY_MAX_SLEEP,
    refine_amounts=True,
    use_cache=True,
    export_format=EXPORT_CASEWARE,
    vendor_map=None,
    accounts_map=None,
    stop_check=None,
):
    if not os.path.isfile(input_path):
        raise ValueError(f"Not a file: {input_path}")
    if export_format not in EXPORT_FORMATS:
        raise ValueError(f"Invalid export format: {export_format}")

    global RATE_LIMITER
    RATE_LIMITER = RateLimiter(rpm)
    vendor_map = load_vendor_map() if vendor_map is None else vendor_map
    accounts_map = load_accounts_map() if accounts_map is None else accounts_map

    output_dir, csv_path, summary_path = _resolve_output_paths_for_file(input_path, output_path)
    rows = _process_file(
        path=input_path,
        source_name=os.path.basename(input_path),
        output_dir=output_dir,
        api_key=api_key,
        model=model,
        max_pages=max_pages,
        zoom=zoom,
        currency=currency,
        max_retries=max_retries,
        retry_backoff=retry_backoff,
        retry_max_sleep=retry_max_sleep,
        refine_amounts_enabled=refine_amounts,
        use_cache=use_cache,
        vendor_map=vendor_map,
        accounts_map=accounts_map,
        stop_check=stop_check,
    )
    if not rows:
        raise RuntimeError("No transactions were extracted.")
    detect_duplicates(rows)
    write_csv(csv_path, rows, export_format=export_format)
    write_summary(summary_path, rows)
    _emit_progress(f"Wrote {csv_path}")
    _emit_progress(f"Wrote {summary_path}")
    return rows, csv_path, summary_path


def main():
    args = parse_args()
    api_key = os.environ.get("OPENROUTER_API_KEY", "").strip()
    try:
        if os.path.isdir(args.input):
            process_directory(
                input_path=args.input,
                api_key=api_key,
                model=args.model,
                output_path=args.output,
                max_pages=args.max_pages,
                zoom=args.zoom,
                currency=args.currency,
                rpm=args.rpm,
                max_retries=args.max_retries,
                retry_backoff=args.retry_backoff,
                retry_max_sleep=args.retry_max_sleep,
                recursive=args.recursive,
                refine_amounts=args.refine_amounts,
                use_cache=args.use_cache,
                export_format=args.export_format,
            )
        else:
            process_single_file(
                input_path=args.input,
                api_key=api_key,
                model=args.model,
                output_path=args.output,
                max_pages=args.max_pages,
                zoom=args.zoom,
                currency=args.currency,
                rpm=args.rpm,
                max_retries=args.max_retries,
                retry_backoff=args.retry_backoff,
                retry_max_sleep=args.retry_max_sleep,
                refine_amounts=args.refine_amounts,
                use_cache=args.use_cache,
                export_format=args.export_format,
            )
    except ProcessingStopped:
        _emit_progress("Stopped by user.", level="WARN")
        sys.exit(2)
    except Exception as exc:
        _emit_progress(f"ERROR: {exc}", level="ERROR")
        sys.exit(1)


if __name__ == "__main__":
    main()
